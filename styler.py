import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.styles import Color, PatternFill, Font, Border, Side
from openpyxl.styles import borders, fills, colors


def text(cell):
    """ The main text describing the sheet

    [text]
    [endtext]
    """
    cell.font = __font_text
    cell.fill = __fill_text


def heading(cell):
    """ A heading in the body of a sheet

    ## A Heading
    """
    cell.font = __font_heading


def label(cell):
    cell.fill = __fill_label


def formula(cell):
    cell.fill = __fill_formula


def value(cell):
    cell.fill = __fill_value


def reference(cell):
    cell.fill = __fill_reference


def obj_inst(cell):
    cell.fill = __fill_obj_inst


def obj_alias(cell):
    cell.fill = __fill_obj_alias


__font_text = Font(name='Calibri',
                   size=11,
                   bold=False,
                   italic=True,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='FF000000')

__font_heading = Font(name='Calibri',
                      size=16,
                      bold=True,
                      italic=False,
                      vertAlign=None,
                      underline='none',
                      strike=False,
                      color='FF000000')

__fill_text = PatternFill(fill_type=fills.FILL_SOLID,
                          start_color=Color('FFFFFFFF'),
                          end_color=Color('FFFFFFFF'))

__fill_label = PatternFill(fill_type=fills.FILL_SOLID,
                           start_color=Color('FFD9D9D9'),
                           end_color=Color('FFD9D9D9'))

__fill_formula = PatternFill(fill_type=fills.FILL_SOLID,
                             start_color=Color('FFD8E4BC'),
                             end_color=Color('FFD8E4BC'))

__fill_value = PatternFill(fill_type=fills.FILL_SOLID,
                           start_color=Color('FFC5D9F1'),
                           end_color=Color('FFC5D9F1'))

__fill_reference = PatternFill(fill_type=fills.FILL_SOLID,
                               start_color=Color('FFDCE6F1'),
                               end_color=Color('FFDCE6F1'))

__fill_obj_inst = PatternFill(fill_type=fills.FILL_SOLID,
                              start_color=Color('FFFCD5B4'),
                              end_color=Color('FFFCD5B4'))

__fill_obj_alias = PatternFill(fill_type=fills.FILL_SOLID,
                               start_color=Color('FFFABF8F'),
                               end_color=Color('FFFABF8F'))
