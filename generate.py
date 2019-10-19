from datetime import datetime

import openpyxl
from openpyxl.utils import coordinate_from_string, column_index_from_string, get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import styler


class WorkBookMaker:
    def __init__(self, fname_markdown, fname_xls):
        with open(fname_markdown) as f:
            self.lines = f.readlines()
        self.fname_xls = fname_xls
        self.wb = openpyxl.Workbook()
        for sheetName in self.wb.sheetnames:
            del self.wb[sheetName]
        self.variables = {}
        self.formulas = []
        self.r = 1
        self.c = 1
        self.ws = None
        self.line_number = -1

    def process(self):
        while self.line_number < (len(self.lines) - 1):
            self.line_number += 1
            line = self.lines[self.line_number]

            # a new sheet
            if line.startswith("# "):
                sheet_name = line[2:]
                self.r = 1
                self.c = 1
                self.wb.create_sheet(sheet_name)
                self.ws = self.wb[sheet_name]
            elif line.startswith("## "):
                value = line[3:]
                self.__cell().value = value
                styler.heading(self.__cell())
                self.r += 1
            elif self.__try_use_cell(line):
                pass
            # a block of markdown text
            elif line.startswith("[text]"):
                self.__process_text()
            elif len(line.strip()) == 0:
                self.r += 1
            # a table of columns, will continue to [endtable_c]
            elif line.startswith("[table_c]"):
                self.__process_table_c()
            # a horizontal create with the object alias to the left
            elif line.startswith("[create_h]"):
                self.__process_create_h()
            else:
                self.__add_values(line)
        self.__update_formulas()

    def save(self):
        """ Save the workbook. Call after process
        """
        self.wb.save(fname_xls)

    def __update_formulas(self):
        """ replace variable names with references
        """
        for f in self.formulas:
            value = f.formula
            for key in self.variables.keys():
                if value.find(key) > -1:
                    value = value.replace(key, self.variables[key])
                    print(f"replacing {key} with {self.variables[key]} : {value} ")
            ws = self.wb[f.sheet]
            ws.cell(f.row, f.col, value)

    def __add_values(self, line):
        """ Handle a comma separated pair and put each of the pair items next to each other

        :param line:
        """
        first = line.find(',')
        if first < 0:
            self.__line_error(line)
        part0 = line[:first]
        part1 = line[first + 1:].replace(' ', '')
        cell = self.__cell()
        cell.value = part0
        styler.label(cell)
        cell = self.__cell(1)
        value_format = "0.00"
        if part1.strip()[0] == "=":
            part1 = part1.strip()
            ind = part1.find('[')
            if ind > - 1:
                value_format = part1[ind + 1:-1]
                part1 = part1[:ind]
            print(f"formula: {part1}")

            self.formulas.append(Formula(self.ws.title, self.r, self.c + 1, part1))
            styler.formula(cell)
            cell.number_format = value_format
        else:
            self.__set_value(cell, part1)
            styler.value(cell)
        self.variables[part0] = f"{get_column_letter(self.c + 1)}{self.r}"
        self.r += 1

    def __set_value(self, cell, str):
        """ determine the number format of the value in str and add it to a cell

        :param cell:
        :param str:
        :return: None
        """
        str = str.strip()
        if str.endswith('%'):
            value = float(str[:-1]) / 100
            cell.value = value
            cell.number_format = '0.00%'
            return
        parts = str.split('-')
        if len(parts) == 3:
            value = v = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
            cell.value = value
            cell.number_format = "YYYY-MM-DD"
            return
        if str.startswith('['):
            i0 = str.find('[')
            i1 = str.find(']')
            list_str = str[i0 + 1:i1]
            value = str[i1 + 1:]
            dv = DataValidation(type="list", formula1=f'"{list_str}"', allow_blank=False)
            dv.add(cell)
            self.ws.add_data_validation(dv)
            cell.value = value
            return
        cell.value = float(str)
        cell.number_format = "0.00"

    def __line_error(self, line):
        print(f"Can not process: {line}")

    def __process_text(self):
        """ Handle a large block of formatted text. Intended to be used to describe the workbook.

        """
        self.line_number += 1
        line = self.lines[self.line_number]
        while not line.startswith("[endtext]"):
            cell = self.ws.cell(row=self.r, column=self.c)
            cell.value = line
            styler.text(cell)
            self.line_number += 1
            self.r += 1
            line = self.lines[self.line_number]
        self.ws.column_dimensions[get_column_letter(self.c)].width = 100

    def __cell(self, col_offset=0):
        """ Get the current cell reference that should be written to next.

        :param col_offset: The offset to the right of the current column
        :return: None
        """
        return self.ws.cell(row=self.r, column=self.c + col_offset)

    def __try_use_cell(self, line):
        if not line.startswith('['):
            return False
        test = line.replace('[', '')
        test = test.replace(']', '')
        test = test.replace(' ', '')
        try:
            xy = coordinate_from_string(test)
            self.r = xy[1]
            self.c = column_index_from_string(xy[0])
            return True
        except:
            return False

    def __process_table_c(self):
        """ Process a table of columns

        """
        self.line_number += 1
        line = self.lines[self.line_number]
        headers = [h.strip() for h in line.split(',')]
        for i in range(len(headers)):
            cell = self.__cell(i)
            cell.value = headers[i]
            styler.label(cell)
        self.r += 1

        self.line_number += 1
        line = self.lines[self.line_number]
        start_row = self.r
        while not line.startswith("[endtable_c]"):
            values = [h.strip() for h in line.split(',')]
            for i in range(len(values)):
                cell = self.__cell(i)
                self.__set_value(cell, values[i])
                styler.value(cell)
            self.line_number += 1
            self.r += 1
            line = self.lines[self.line_number]
        end_row = self.r - 1

        for i in range(len(headers)):
            col = get_column_letter(self.c + i)
            self.variables[headers[i]] = f"{col}{start_row}:{col}{end_row}"

    def __range_name(self, r, c):
        """ Get the range name from a row and column

        :param r:
        :param c:
        :return: The range name in A1 format.
        """
        return f"{get_column_letter(c)}{r}"

    def __process_create_h(self):
        """ Horizontal object creation with alias to left.

        :return: None
        """

        self.line_number += 1
        parts = self.lines[self.line_number].strip().split(',', 1)
        self.variables[parts[0]] = self.__range_name(self.r, self.c + 1)
        cell = self.__cell()
        cell.value = parts[0]
        styler.obj_alias(cell)
        create_formula = parts[1].replace('**', self.__range_name(self.r, self.c))
        self.formulas.append(Formula(self.ws.title, self.r, self.c + 1, create_formula))
        cell = self.__cell(1)
        styler.obj_inst(cell)
        self.line_number += 1


class Formula:
    """ The data required to store a formula and its location in the sheet.
    """

    def __init__(self, sheet, row, col, formula):
        self.sheet = sheet
        self.row = row
        self.col = col
        self.formula = formula


fname_markdown = "C:/Dev/python/xls_markdown/tiny_example.md"
fname_xls = "C:/Dev/python/xls_markdown/tiny_example.xlsx"
wbm = WorkBookMaker(fname_markdown, fname_xls)
wbm.process()
wbm.save()
